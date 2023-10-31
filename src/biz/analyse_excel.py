from openpyxl import load_workbook
import asyncio
import os


class AnalyseExcel:
    def __init__(self, file_path: str) -> None:
        self.file_path = file_path

    async def __aenter__(self):
        await asyncio.sleep(0.1)
        self.wb = load_workbook(self.file_path)
        return self

    async def __aexit__(self, exc_type, exc_val, exc_tb):
        self.wb.close()

    async def get_data_num(self) -> int:
        ws = self.wb.active
        data_row_num = -1
        for row in ws.rows:
            for cell in row:
                if isinstance(cell.value, str) and 'SCM' in cell.value:
                    data_row_num = cell.row - 1
                    break
            if data_row_num != -1:
                break
        return ws.max_row - data_row_num

    async def get_pic_num(self) -> int:
        ws = self.wb.active

        num = 0
        for row in ws.rows:
            for cell in row:
                if isinstance(cell.value, str) and 'SCM' in cell.value:
                    num += 1
        return num

    async def get_pic_downloaded_num(self) -> int:
        count = 0
        for root, dirs, files in os.walk(os.path.dirname(self.file_path)):
            for file in files:
                if file.endswith(".jpg") or file.endswith(".jpeg") or file.endswith(".png"):
                    count += 1
        return count

    async def get_pic_compressed_num(self) -> int:
        count = 0
        for root, dirs, files in os.walk(os.path.dirname(self.file_path)):
            for file in files:
                if file.endswith(".compress"):
                    count += 1
        return count
