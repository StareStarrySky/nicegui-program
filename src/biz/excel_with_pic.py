from copy import copy

import requests
import os
import asyncio

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.workbook import Workbook
from openpyxl.drawing.image import Image
from io import BytesIO
from openpyxl.utils import get_column_letter
from PIL import Image as PILImage

from exception.biz_except import BizExcept


class ExcelWithPic:

    def __init__(self, excel_path: str, *,
                 url_prefix: str = 'http://10.0.0.100/', pic_len: int = 10,
                 pic_width: int = 0, pic_height: int = 0, compress_rate: float = 0.8,
                 res_num: int = 4) -> None:
        """ExcelWithPic

        替换Excel里的图片地址

        :param excel_path: 源文件路径
        :param url_prefix: 图片地址前缀
        :param pic_len: 每张图片压缩至(kb)
        :param compress_rate: 每次压缩比例
        :param res_num: 输出文件个数
        """
        self.__window = None

        if len(excel_path) == 0:
            raise BizExcept('文件路径不能为空')
        self.excel_path = excel_path
        self.root_path = os.path.dirname(excel_path) + '\\'
        if len(url_prefix) == 0:
            raise BizExcept('图片地址前缀不能为空')
        self.url_prefix = url_prefix
        self.pic_len = pic_len
        self.pic_width = pic_width
        self.pic_height = pic_height
        self.compress_rate = compress_rate
        self.res_num = res_num

        self.complete_num = 0

    async def __aenter__(self):
        return self

    async def __aexit__(self, exc_type, exc_val, exc_tb):
        pass

    def set_window(self, window):
        if self.__window is None:
            self.__window = window

    async def run(self):
        await asyncio.sleep(0.1)
        if self.__window.data_num == 0:
            raise BizExcept('请先分析文件，获取总数')

        wb = load_workbook(self.excel_path)

        try:
            await self.transfer_pic(wb)
            self.__window.success = True
        finally:
            wb.close()

    async def transfer_pic(self, wb):
        ws = wb.active
        if self.__window.res_num < self.__window.data_num:
            res_file_num = self.__window.data_num // self.__window.res_num + 1
            for i in range(res_file_num):
                new_wb = Workbook()
                try:
                    new_ws = new_wb.active
                    await self.transfer_pic_row(ws, i + 1, self.__window.res_num, new_ws)
                    new_wb.save(self.excel_path + '_' + str(i) + '.xlsx')
                finally:
                    new_wb.close()
        else:
            new_wb = Workbook()
            try:
                new_ws = new_wb.active
                await self.transfer_pic_row(ws, 1, self.__window.res_num, new_ws)
                new_wb.save(self.excel_path + '_new.xlsx')
            finally:
                new_wb.close()

    async def transfer_pic_row(self, ws, page, page_num, new_ws):
        data_row_num = -1
        for row in ws.rows:
            for cell in row:
                if isinstance(cell.value, str) and 'SCM' in cell.value:
                    data_row_num = cell.row - 1
                    break
            if data_row_num != -1:
                break
        for i, row in enumerate(ws.rows):
            if i < data_row_num:
                for cell in row:
                    new_cell = new_ws.cell(cell.row, cell.column)
                    self.transfer_pic_copy(ws, cell, new_cell)
            else:
                break
        for i, row in enumerate(ws.rows):
            if i < data_row_num:
                continue
            if (page - 1) * page_num + data_row_num <= i < page * page_num + data_row_num:
                self.transfer_pic_cell(ws, row, (page - 1) * page_num, new_ws)

                await self.update_process()

    async def update_process(self):
        await asyncio.sleep(0)
        self.complete_num += 1
        self.__window.data_rate = round(self.complete_num / self.__window.data_num, 2)
        self.__window.data_rate_str = f'{self.complete_num} / {self.__window.data_num}'

    def transfer_pic_copy(self, ws, cell, new_cell):
        if isinstance(cell, MergedCell):
            return
        new_cell.data_type = ws[cell.coordinate].data_type
        if ws[cell.coordinate].has_style:
            new_cell.font = copy(ws[cell.coordinate].font)
            new_cell.border = copy(ws[cell.coordinate].border)
            new_cell.fill = copy(ws[cell.coordinate].fill)
            new_cell.number_format = copy(ws[cell.coordinate].number_format)
            new_cell.alignment = copy(ws[cell.coordinate].alignment)
        new_cell.value = ws[cell.coordinate].value

    def transfer_pic_cell(self, ws, row, offset, new_ws):
        for cell in row:
            new_cell = new_ws.cell(cell.row - offset, cell.column)

            self.transfer_pic_copy(ws, cell, new_cell)

            if isinstance(new_cell.value, str) and 'SCM' in new_cell.value:
                pic = self.get_pic(new_cell.value)
                if len(pic) == 0:
                    continue
                with PILImage.open(BytesIO(pic)) as im:
                    cell_size = im.width * 0.125, im.height * 0.75

                ws_width = new_ws.column_dimensions[get_column_letter(new_cell.column)].width
                if ws_width is None:
                    ws_width = 8.38

                if ws_width < cell_size[0]:
                    new_ws.column_dimensions[get_column_letter(new_cell.column)].width = cell_size[0]

                ws_height = new_ws.row_dimensions[new_cell.row].height
                if ws_height is None:
                    ws_height = 14.25

                if ws_height < cell_size[1]:
                    new_ws.row_dimensions[new_cell.row].height = cell_size[1]

                img = Image(BytesIO(pic))
                new_ws.add_image(img, new_cell.coordinate)

    def download_pic(self, pic_path):
        headers = {'Referer': self.url_prefix}
        try:
            res = requests.get(f'{self.url_prefix}tlmpUpload/{pic_path}', headers=headers)
            if res.status_code == 200:
                return res.content
            else:
                return b''
        except requests.exceptions.RequestException:
            return b''

    def compress(self, pic_src):
        pic_tar = pic_src
        tar = len(pic_tar)
        if tar <= self.pic_len * 1024:
            return pic_src
        while tar > self.pic_len * 1024:
            with PILImage.open(BytesIO(pic_tar)) as im:
                (width, height) = (int(im.width * self.compress_rate), int(im.height * self.compress_rate))
                if self.pic_width > 0:
                    while width > self.pic_width:
                        (width, height) = (int(width * self.compress_rate), int(height * self.compress_rate))
                if self.pic_height > 0:
                    while height > self.pic_height:
                        (width, height) = (int(width * self.compress_rate), int(height * self.compress_rate))
                im_com = im.resize((width, height))
            pic_tar_bytes = BytesIO()
            try:
                im_com.save(pic_tar_bytes, 'JPEG')
            except OSError:
                im_com.save(pic_tar_bytes, 'PNG')
            pic_tar = pic_tar_bytes.getvalue()
            tar = len(pic_tar)
        return pic_tar

    def get_pic(self, pic_path):
        local_path = self.root_path + pic_path
        if os.path.exists(local_path + '.compress'):
            with open(local_path + '.compress', 'rb') as pic_obj:
                return pic_obj.read()
        elif os.path.exists(local_path):
            with open(local_path, 'rb') as pic_obj:
                pic_bytes = pic_obj.read()
            pic_com_bytes = self.compress(pic_bytes)
            if len(pic_com_bytes) != len(pic_bytes):
                with open(local_path + '.compress', 'wb') as pic_compress:
                    pic_compress.write(pic_com_bytes)
            return pic_com_bytes
        else:
            pic_dir = os.path.dirname(local_path)
            if not os.path.exists(pic_dir):
                os.makedirs(pic_dir, exist_ok=True)
            pic_bytes = self.download_pic(pic_path)
            if len(pic_bytes) > 0:
                with open(local_path, 'wb') as pic_obj:
                    pic_obj.write(pic_bytes)
                pic_com_bytes = self.compress(pic_bytes)
                if len(pic_com_bytes) != len(pic_bytes):
                    with open(local_path + '.compress', 'wb') as pic_compress:
                        pic_compress.write(pic_com_bytes)
                return pic_com_bytes
            else:
                return b''
